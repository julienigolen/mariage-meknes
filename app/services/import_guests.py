"""Logique de parsing/écriture de l'import Excel — extraite de scripts/import_guests.py
le 2026-07-28 (construction admin) pour être appelée à la fois par le script CLI et par
l'upload web (app/routers/admin.py). Une seule source de vérité pour les deux chemins.

Format aligné sur la table admin (Patron 2026-07-29, « aligne le format du fichier
import/export sur le dernier format du tableau ») : **une ligne par foyer**, pas par
personne — la logique de regroupement par libellé Famille a disparu, chaque ligne porte
déjà son invité principal et son invité secondaire optionnel (2 personnes maximum par
foyer, même contrainte que la table). Colonnes attendues (en-têtes en première ligne,
insensibles à la casse/aux espaces) :

    Invité principal   (nom, obligatoire)
    Tél. principal      (avec indicatif, déjà complet — §5.9 proposition_produit.md)
    Invité secondaire  (optionnel — nom du 2e invité du foyer)
    Tél. secondaire     (obligatoire SI Invité secondaire est rempli, et vice-versa)
    Famille             (optionnel — libellé libre, purement descriptif désormais)
    Effectif théorique  (optionnel — défaut 1, porté par le foyer via l'invité principal)
    Origine             (optionnel — "fr" ou "ma", défaut "fr" — retro-compat, plus affiché
                          dans la table)

Rupture de format (2026-07-29) : les fichiers exportés/importés avant cette date (une
ligne par personne, colonnes Nom Prénom/Téléphone/Famille) ne sont plus compatibles —
assumé, aucune donnée réelle n'est encore en production à cette date.

Doublons de téléphone (2026-07-28, ajouté pour l'upload web) : un numéro déjà en base
— import précédent, OU foyer auto-ajouté via /rsvp/join (§ « invité non répertorié »),
OU doublon à l'intérieur même du fichier — est ignoré avec un avertissement plutôt que
de faire échouer toute la transaction. Si seul le secondaire est en doublon, le foyer
et son principal sont quand même créés (seul le secondaire est ignoré).
"""
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, BinaryIO

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Household, HouseholdMember
from app.phone import normalize_phone

EXPORT_HEADERS = [
    "Invité principal", "Tél. principal", "Famille", "Effectif théorique",
    "Statut", "Présence", "Adultes", "Enfants", "Hôtel",
    "Invité secondaire", "Tél. secondaire", "Allergies", "Commentaire allergies",
]

HEADER_ALIASES = {
    "nom_principal": ["invite principal", "invité principal", "nom principal", "principal"],
    "tel_principal": ["tel principal", "tél principal", "telephone principal", "téléphone principal"],
    "nom_secondaire": ["invite secondaire", "invité secondaire", "nom secondaire", "secondaire"],
    "tel_secondaire": ["tel secondaire", "tél secondaire", "telephone secondaire", "téléphone secondaire"],
    "famille": ["famille"],
    "origine": ["origine"],
    "effectif_theorique": ["effectif theorique", "effectif théorique", "nombre theorique", "nombre théorique"],
}


def _clean(s: Any) -> str:
    return str(s).strip() if s is not None else ""


def _find_columns(headers: list[str]) -> dict[str, int]:
    # .replace(".", "") : les en-têtes exportés portent des abréviations ("Tél.") que
    # les alias n'ont pas besoin de dupliquer avec et sans point (bug relevé le
    # 2026-07-29 : "Tél. principal" ne matchait pas l'alias "tel principal").
    normalized = [_clean(h).lower().replace(".", "") for h in headers]
    found: dict[str, int] = {}
    for field_name, aliases in HEADER_ALIASES.items():
        for i, h in enumerate(normalized):
            if h in aliases:
                found[field_name] = i
                break
    return found


def _deduce_langue(phone: str) -> str:
    if phone.startswith("+33"):
        return "fr"
    if phone.startswith("+212"):
        return "ar"
    return "fr"


@dataclass
class ImportResult:
    header_error: str | None = None
    n_lignes: int = 0
    erreurs: list[str] = field(default_factory=list)
    doublons: list[str] = field(default_factory=list)
    # Un foyer par ligne valide : {"famille": str | None, "members": [dict, ...]} —
    # 1 ou 2 dicts prêts pour HouseholdMember(**dict).
    foyers: list[dict] = field(default_factory=list)

    @property
    def n_personnes(self) -> int:
        return sum(len(f["members"]) for f in self.foyers)

    @property
    def n_foyers(self) -> int:
        return len(self.foyers)


def parse_workbook(file: BinaryIO, existing_phones: set[str]) -> ImportResult:
    """Lit le classeur, une ligne = un foyer — n'écrit rien en base.
    `existing_phones` : numéros déjà connus (import précédent + auto-ajouts RSVP), pour
    détecter les doublons avant écriture plutôt qu'à l'échec de la contrainte unique."""
    result = ImportResult()
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result.header_error = "Fichier vide."
        return result

    headers = list(rows[0])
    cols = _find_columns(headers)

    if "nom_principal" not in cols:
        result.header_error = f"Colonne « Invité principal » introuvable. En-têtes lus : {headers}"
        return result
    if "tel_principal" not in cols:
        result.header_error = f"Colonne « Tél. principal » introuvable. En-têtes lus : {headers}"
        return result

    seen_phones = set(existing_phones)

    for row_idx, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None for c in row):
            continue
        result.n_lignes += 1

        nom_principal = _clean(row[cols["nom_principal"]])
        phone_principal_raw = _clean(row[cols["tel_principal"]])
        famille = _clean(row[cols["famille"]]) if "famille" in cols else ""
        origine = (_clean(row[cols["origine"]]) or "fr").lower() if "origine" in cols else "fr"
        origine = origine if origine in ("fr", "ma") else "fr"
        effectif_raw = _clean(row[cols["effectif_theorique"]]) if "effectif_theorique" in cols else ""
        try:
            effectif_theorique = int(float(effectif_raw)) if effectif_raw else 1
            if effectif_theorique < 1:
                effectif_theorique = 1
        except ValueError:
            effectif_theorique = 1

        if not nom_principal or not phone_principal_raw:
            result.erreurs.append(f"ligne {row_idx} : invité principal (nom ou téléphone) manquant, ligne ignorée")
            continue

        phone_principal = normalize_phone(phone_principal_raw)
        if not phone_principal.startswith("+"):
            result.erreurs.append(
                f"ligne {row_idx} ({nom_principal}) : numéro principal « {phone_principal_raw} » "
                "sans indicatif reconnu (+..), importé tel quel"
            )
        if phone_principal in seen_phones:
            result.doublons.append(f"ligne {row_idx} ({nom_principal}) : numéro principal {phone_principal} déjà connu, foyer ignoré")
            continue
        seen_phones.add(phone_principal)

        members = [{
            "nom_prenom": nom_principal,
            "phone": phone_principal,
            "origine": origine,
            "langue": _deduce_langue(phone_principal),
            "import_source": f"ligne {row_idx}",
            "effectif_theorique": effectif_theorique,
        }]

        nom_secondaire = _clean(row[cols["nom_secondaire"]]) if "nom_secondaire" in cols else ""
        phone_secondaire_raw = _clean(row[cols["tel_secondaire"]]) if "tel_secondaire" in cols else ""

        if nom_secondaire and phone_secondaire_raw:
            phone_secondaire = normalize_phone(phone_secondaire_raw)
            if not phone_secondaire.startswith("+"):
                result.erreurs.append(
                    f"ligne {row_idx} ({nom_secondaire}) : numéro secondaire « {phone_secondaire_raw} » "
                    "sans indicatif reconnu (+..), importé tel quel"
                )
            if phone_secondaire in seen_phones:
                result.doublons.append(
                    f"ligne {row_idx} ({nom_secondaire}) : numéro secondaire {phone_secondaire} déjà connu, invité secondaire ignoré (foyer conservé)"
                )
            else:
                seen_phones.add(phone_secondaire)
                members.append({
                    "nom_prenom": nom_secondaire,
                    "phone": phone_secondaire,
                    "origine": origine,
                    "langue": _deduce_langue(phone_secondaire),
                    "import_source": f"ligne {row_idx}",
                    "effectif_theorique": 1,  # l'effectif théorique du foyer est porté par le principal seul
                })
        elif nom_secondaire or phone_secondaire_raw:
            result.erreurs.append(
                f"ligne {row_idx} ({nom_principal}) : invité secondaire à moitié renseigné (nom ou téléphone manquant), ignoré"
            )

        result.foyers.append({"famille": famille or None, "members": members})

    return result


def write_import(db: Session, result: ImportResult) -> None:
    """Écrit le résultat d'un parse_workbook() réussi. Transactionnel : tout ou rien."""
    for foyer in result.foyers:
        household = Household(import_famille_label=foyer["famille"])
        db.add(household)
        db.flush()
        for m in foyer["members"]:
            db.add(HouseholdMember(household_id=household.id, **m))
    db.commit()


def existing_phones(db: Session) -> set[str]:
    return set(db.execute(select(HouseholdMember.phone)).scalars().all())


def export_workbook(households: list[Household]) -> BytesIO:
    """Export « au format d'import » (Patron 2026-07-28, réaligné le 2026-07-29 sur la
    table admin) : mêmes en-têtes que parse_workbook() attend pour le principal/
    secondaire/famille/effectif — réimportable tel quel. Statut/Présence/Adultes/
    Enfants/Hôtel/Allergies sont des colonnes de LECTURE SEULE (état RSVP courant,
    reflète la table) : ignorées par parse_workbook (hors HEADER_ALIASES), le RSVP ne
    se réimporte jamais depuis Excel, seul le site l'écrit."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invités"
    ws.append(EXPORT_HEADERS)
    for hh in households:
        principal = hh.members[0] if hh.members else None
        secondaire = hh.members[1] if len(hh.members) > 1 else None
        r = hh.rsvp
        if r is None:
            statut = "Invité"
        elif r.presence:
            statut = "Accepté"
        else:
            statut = "Refusé"
        ws.append([
            principal.nom_prenom if principal else "",
            principal.phone if principal else "",
            hh.import_famille_label or "",
            principal.effectif_theorique if principal else 1,
            statut,
            "Oui" if (r and r.presence) else "Non",
            r.nb_adultes if r else 0,
            r.nb_enfants if r else 0,
            "Oui" if (r and r.besoin_hotel) else "Non",
            secondaire.nom_prenom if secondaire else "",
            secondaire.phone if secondaire else "",
            "Oui" if (r and r.allergies_bool) else "Non",
            r.allergies_texte if (r and r.allergies_texte) else "",
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
